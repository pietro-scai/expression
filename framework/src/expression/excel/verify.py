"""Round-trip verification (PRD §9.2 step 6).

Open the produced ``.xlsx``, evaluate the formulas, diff each row/period cell
against the in-memory solve. Default tolerance is ``1e-9`` relative.

Strategy:
1. If the optional `formulas` library is installed, use it to evaluate the
   workbook (full Excel-formula semantics).
2. Otherwise, fall back to a pure-Python evaluator that handles the formula
   dialect we *generate* in :mod:`.export` (no inputs from third-party
   workbooks, just our own emitted formulas). This keeps the round-trip path
   working with zero extra deps for Phase 2 acceptance.

The fallback is intentionally narrow: it covers the operators / function
calls the export side emits (``+`` ``-`` ``*`` ``/`` ``^``, comparisons,
``IF``, ``AND``, ``OR``, ``NOT``, ``MOD``, ``INT``, ``SUM``, named ranges,
direct cell refs, range refs).
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from ..core import Model


@dataclass
class Mismatch:
    sheet: str
    cell: str
    label: str
    period: Any
    expected: Any
    actual: Any


@dataclass
class VerifyResult:
    tolerance: float = 1e-9
    checked: int = 0
    mismatches: list[Mismatch] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.mismatches

    def first_mismatches(self, n: int = 10) -> list[Mismatch]:
        return self.mismatches[:n]


def verify(model: Model, path: str | Path, tolerance: float = 1e-9) -> VerifyResult:
    """Compare the workbook at ``path`` against the in-memory ``model`` solve.

    The model is assumed solved already (else we solve it).
    """
    if not getattr(model, "_cells", None):
        model.solve()

    p = Path(path)
    wb = load_workbook(p, data_only=False)
    evaluator = _make_evaluator(wb, p)

    result = VerifyResult(tolerance=tolerance)

    if "model" not in wb.sheetnames:
        raise ValueError(f"{p}: missing 'model' sheet")
    ws = wb["model"]

    period_col: dict[Any, str] = {}
    for col_idx in range(2, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is None:
            continue
        period_col[cell.value] = get_column_letter(col_idx)

    for row_idx in range(2, ws.max_row + 1):
        label_cell = ws.cell(row=row_idx, column=1)
        label = label_cell.value
        if not label:
            continue
        for period, col_letter in period_col.items():
            cell_addr = f"{col_letter}{row_idx}"
            expected = _maybe_get(model, str(label), period)
            if expected is None:
                continue
            actual = evaluator(ws.title, cell_addr)
            if not _close(expected, actual, tolerance):
                result.mismatches.append(
                    Mismatch(
                        sheet=ws.title,
                        cell=cell_addr,
                        label=str(label),
                        period=period,
                        expected=expected,
                        actual=actual,
                    )
                )
            result.checked += 1
    return result


def _maybe_get(model: Model, label: str, period: Any) -> Any:
    if model.has_cell(label, period):
        return model.cell(label, period)
    return None


def _close(a: Any, b: Any, tol: float) -> bool:
    if a is None or b is None:
        return a == b
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return math.isclose(float(a), float(b), rel_tol=tol, abs_tol=tol)
    return a == b


# ---------------------------------------------------------------------------
# Evaluator: prefer `formulas`; fall back to a small built-in evaluator.
# ---------------------------------------------------------------------------


def _make_evaluator(wb: Any, path: Path):
    try:
        return _formulas_lib_evaluator(path)
    except Exception:
        return _fallback_evaluator(wb)


def _formulas_lib_evaluator(path: Path):
    """Use the optional ``formulas`` library if installed."""
    import formulas  # type: ignore[import-not-found]

    xl = formulas.ExcelModel().loads(str(path)).finish()
    sol = xl.calculate()

    def evaluate(sheet: str, cell: str) -> Any:
        key = f"'[{path.name}]{sheet}'!{cell}".upper()
        if key in sol:
            v = sol[key].value
            return _unbox(v)
        # try without quotes
        key = f"[{path.name}]{sheet}!{cell}".upper()
        if key in sol:
            return _unbox(sol[key].value)
        return None

    return evaluate


def _unbox(v: Any) -> Any:
    """`formulas` wraps results in arrays / Sympy types — unwrap to plain Python."""
    try:
        # most numeric scalars come out as ndarray-like with .item()
        return v.item() if hasattr(v, "item") else v
    except Exception:
        return v


def _fallback_evaluator(wb: Any):
    """Tiny formula evaluator for the dialect we emit.

    Recognises absolute / relative refs, named ranges, ranges (``A1:Z1``),
    numbers, strings, ``+ - * / ^``, comparisons, and a few funcs.
    """
    cache: dict[tuple[str, str], Any] = {}
    in_progress: set[tuple[str, str]] = set()

    # name -> (sheet, cell)
    named: dict[str, tuple[str, str]] = {}
    for name, dn in wb.defined_names.items():
        # attr_text like "'globals'!$B$2"
        ref = dn.attr_text
        m = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", ref)
        if m:
            named[name] = (m.group(1), f"{m.group(2)}{m.group(3)}")

    def get_cell(sheet: str, cell: str) -> Any:
        key = (sheet, cell.replace("$", ""))
        if key in cache:
            return cache[key]
        if key in in_progress:
            raise RuntimeError(f"cycle through {key}")
        in_progress.add(key)
        try:
            value = _evaluate_cell(sheet, key[1])
            cache[key] = value
            return value
        finally:
            in_progress.discard(key)

    def _evaluate_cell(sheet: str, cell: str) -> Any:
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        c = ws[cell]
        v = c.value
        if isinstance(v, str) and v.startswith("="):
            return _eval_expr(v[1:], sheet)
        return v

    def _eval_expr(expr: str, sheet: str) -> Any:
        return _Parser(expr, sheet, get_cell, named, wb).parse()

    return get_cell


# ---------------------------------------------------------------------------
# Mini formula parser (Pratt-ish, ad-hoc) for the dialect we emit.
# ---------------------------------------------------------------------------


_TOKEN_RE = re.compile(
    r"""
    \s*(?:
        (?P<num>-?\d+(?:\.\d+)?(?:[eE][+\-]?\d+)?)
      | (?P<str>"(?:[^"]|"")*")
      | (?P<range>\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)
      | (?P<ref>\$?[A-Z]+\$?\d+)
      | (?P<ident>[A-Za-z_][A-Za-z_0-9]*)
      | (?P<op><=|>=|<>|[+\-*/^()<>=,])
    )
    """,
    re.VERBOSE,
)


class _Parser:
    def __init__(
        self,
        expr: str,
        sheet: str,
        get_cell,
        named: dict[str, tuple[str, str]],
        wb: Any,
    ):
        self.expr = expr
        self.sheet = sheet
        self.get_cell = get_cell
        self.named = named
        self.wb = wb
        self.tokens = self._tokenize(expr)
        self.pos = 0

    def _tokenize(self, expr: str) -> list[tuple[str, str]]:
        toks: list[tuple[str, str]] = []
        pos = 0
        while pos < len(expr):
            if expr[pos].isspace():
                pos += 1
                continue
            m = _TOKEN_RE.match(expr, pos)
            if not m:
                raise ValueError(f"bad token at pos {pos} in {expr!r}")
            for k, v in m.groupdict().items():
                if v is not None:
                    toks.append((k, v))
                    break
            pos = m.end()
        return toks

    def parse(self) -> Any:
        v = self._parse_compare()
        if self.pos != len(self.tokens):
            raise ValueError(f"trailing tokens: {self.tokens[self.pos :]}")
        return v

    def _peek(self) -> tuple[str, str] | None:
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def _eat(self, kind: str | None = None, value: str | None = None) -> tuple[str, str]:
        tok = self.tokens[self.pos]
        if kind and tok[0] != kind:
            raise ValueError(f"expected {kind}, got {tok}")
        if value and tok[1] != value:
            raise ValueError(f"expected {value!r}, got {tok}")
        self.pos += 1
        return tok

    def _parse_compare(self) -> Any:
        left = self._parse_add()
        tok = self._peek()
        while tok and tok[0] == "op" and tok[1] in ("<", "<=", ">", ">=", "=", "<>"):
            self._eat()
            right = self._parse_add()
            left = _compare(tok[1], left, right)
            tok = self._peek()
        return left

    def _parse_add(self) -> Any:
        left = self._parse_mul()
        tok = self._peek()
        while tok and tok[0] == "op" and tok[1] in ("+", "-"):
            self._eat()
            right = self._parse_mul()
            left = (left + right) if tok[1] == "+" else (left - right)
            tok = self._peek()
        return left

    def _parse_mul(self) -> Any:
        left = self._parse_pow()
        tok = self._peek()
        while tok and tok[0] == "op" and tok[1] in ("*", "/"):
            self._eat()
            right = self._parse_pow()
            left = (left * right) if tok[1] == "*" else (left / right)
            tok = self._peek()
        return left

    def _parse_pow(self) -> Any:
        left = self._parse_unary()
        tok = self._peek()
        if tok and tok[0] == "op" and tok[1] == "^":
            self._eat()
            right = self._parse_unary()
            return left**right
        return left

    def _parse_unary(self) -> Any:
        tok = self._peek()
        if tok and tok[0] == "op" and tok[1] == "-":
            self._eat()
            return -self._parse_unary()
        if tok and tok[0] == "op" and tok[1] == "+":
            self._eat()
            return self._parse_unary()
        return self._parse_atom()

    def _parse_atom(self) -> Any:
        tok = self._peek()
        if tok is None:
            raise ValueError("unexpected end of expression")
        kind, val = tok
        if kind == "num":
            self._eat()
            return float(val) if "." in val or "e" in val or "E" in val else int(val)
        if kind == "str":
            self._eat()
            return val[1:-1].replace('""', '"')
        if kind == "range":
            self._eat()
            return self._eval_range(val)
        if kind == "ref":
            self._eat()
            return self.get_cell(self.sheet, val)
        if kind == "ident":
            self._eat()
            # is it a named range?
            upper = val.upper()
            if upper == "TRUE":
                return True
            if upper == "FALSE":
                return False
            # function call?
            nxt = self._peek()
            if nxt and nxt[0] == "op" and nxt[1] == "(":
                return self._call_function(val)
            if val in self.named:
                sheet, cell = self.named[val]
                return self.get_cell(sheet, cell)
            raise ValueError(f"unknown name {val!r}")
        if kind == "op" and val == "(":
            self._eat()
            v = self._parse_compare()
            self._eat("op", ")")
            return v
        raise ValueError(f"unexpected token {tok}")

    def _call_function(self, name: str) -> Any:
        self._eat("op", "(")
        args: list[Any] = []
        if self._peek() != ("op", ")"):
            args.append(self._parse_compare())
            while self._peek() == ("op", ","):
                self._eat()
                args.append(self._parse_compare())
        self._eat("op", ")")
        return _call_excel_func(name.upper(), args)

    def _eval_range(self, ref: str) -> list[Any]:
        a, b = ref.split(":")
        ma = re.match(r"\$?([A-Z]+)\$?(\d+)", a)
        mb = re.match(r"\$?([A-Z]+)\$?(\d+)", b)
        if not (ma and mb):
            raise ValueError(f"bad range {ref!r}")
        c1, r1 = column_index_from_string(ma.group(1)), int(ma.group(2))
        c2, r2 = column_index_from_string(mb.group(1)), int(mb.group(2))
        out: list[Any] = []
        for cc in range(c1, c2 + 1):
            for rr in range(r1, r2 + 1):
                out.append(self.get_cell(self.sheet, f"{get_column_letter(cc)}{rr}"))
        return out


def _compare(op: str, a: Any, b: Any) -> bool:
    if op == "=":
        return a == b
    if op == "<>":
        return a != b
    if op == "<":
        return a < b
    if op == "<=":
        return a <= b
    if op == ">":
        return a > b
    if op == ">=":
        return a >= b
    raise ValueError(f"bad compare op {op}")


def _flatten(args: list[Any]) -> list[Any]:
    out: list[Any] = []
    for a in args:
        if isinstance(a, list):
            out.extend(_flatten(a))
        else:
            out.append(a)
    return out


def _call_excel_func(name: str, args: list[Any]) -> Any:
    if name == "IF":
        cond, when_true, when_false = args
        return when_true if cond else when_false
    if name == "AND":
        return all(_flatten(args))
    if name == "OR":
        return any(_flatten(args))
    if name == "NOT":
        return not args[0]
    if name == "MOD":
        return args[0] % args[1]
    if name == "INT":
        return int(args[0])
    if name == "SUM":
        return sum(x for x in _flatten(args) if isinstance(x, (int, float)))
    if name == "AVERAGE":
        flat = [x for x in _flatten(args) if isinstance(x, (int, float))]
        return sum(flat) / len(flat) if flat else 0
    if name == "MIN":
        return min(x for x in _flatten(args) if x is not None)
    if name == "MAX":
        return max(x for x in _flatten(args) if x is not None)
    raise ValueError(f"unknown Excel function {name}")
