"""Layout helpers: map model rows + periods to Excel cells (PRD §9.2).

The layout for Phase 2:

- Sheet ``model``:
    - Row 1: header. Cell A1 = "(label)"; B1..N1 = period labels (years).
    - Row 2..K+1: one row per ``@row`` in topological order.
        - Column A = row name (the label).
        - Columns B..N = the formula or value for each period.
- Sheet ``globals``:
    - One row per ``glob``: A=name, B=value. B is a named range whose name
      is the glob name (so formulas can write ``=growth_rate``).
- Sheet ``scalars``:
    - One row per ``@scalar``: A=name, B=value (a precomputed constant in
      Phase 2 — full Excel-formula scalars are Phase 3+).
- Sheets ``ds_<name>`` (hidden) for each ``dataset.csv`` attribute, holding
  the dataset rows as a header-row + data-rows table.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter


@dataclass
class Layout:
    """Resolved layout for a single model export run."""

    # row_name -> 1-based row index on the "model" sheet
    row_index: dict[str, int] = field(default_factory=dict)
    # period -> 1-based column index on the "model" sheet (B=2, C=3, ...)
    col_index: dict[object, int] = field(default_factory=dict)
    # set of row names (used by the formula-emitter to recognize references)
    row_names: set[str] = field(default_factory=set)
    # set of glob names (referenced as named ranges in formulas)
    glob_names: set[str] = field(default_factory=set)
    # set of scalar names (referenced via the scalars sheet, named ranges)
    scalar_names: set[str] = field(default_factory=set)
    # name of the sheet holding the row table (default "model")
    sheet: str = "model"
    # row 1 is the header; first data row is row 2
    header_row: int = 1
    # column A is labels; first data column is B (=2)
    label_col: int = 1

    def cell_ref(self, row_name: str, period: object) -> str:
        """Return an absolute Excel cell ref like ``$B$2`` for (row, period)."""
        if row_name not in self.row_index:
            raise KeyError(f"row {row_name!r} not in layout")
        if period not in self.col_index:
            raise KeyError(f"period {period!r} not in layout")
        col = get_column_letter(self.col_index[period])
        return f"{col}{self.row_index[row_name]}"

    def relative_cell_ref(self, row_name: str, period: object) -> str:
        """Like :meth:`cell_ref` but without ``$`` (for in-row formulas)."""
        col = get_column_letter(self.col_index[period])
        return f"{col}{self.row_index[row_name]}"
