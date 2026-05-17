"""Render a solved :class:`~model.core.Model` to an ``.xlsx`` workbook.

Layout (one row per ``@row``, one column per period — see ``_layout``).

For each cell we emit either a *formula* (mirroring the Python expression) or
a *literal value* (when the expression can't be cleanly translated yet — see
:class:`._formula.UnsupportedFormula`). Round-trip values are preserved
either way; the verifier diffs values, not formulas.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from ..core import Model
from ._formula import UnsupportedFormula, row_formula_for_period
from ._layout import Layout

_NUMBER_FMT = "#,##0.00"
_HEADER_FMT_KEY = "header"


def export(model: Model, path: str | Path) -> Path:
    """Solve ``model`` and write to ``path`` as ``.xlsx``.

    Returns the resolved ``Path``.
    """
    if not getattr(model, "_cells", None):
        model.solve()

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # remove the default sheet — we'll create our own
    default = wb.active
    if default is not None:
        wb.remove(default)

    layout = _build_layout(model)
    _write_globals_sheet(wb, model, layout)
    _write_scalars_sheet(wb, model, layout)
    _write_datasets(wb, model)
    _write_model_sheet(wb, model, layout)

    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Layout
# ---------------------------------------------------------------------------


def _build_layout(model: Model) -> Layout:
    layout = Layout()
    rows = model._rows  # pyright: ignore[reportPrivateUsage]
    layout.row_names = set(rows)
    layout.glob_names = set(model._globs)  # pyright: ignore[reportPrivateUsage]
    layout.scalar_names = set(model._scalars)  # pyright: ignore[reportPrivateUsage]

    # Period columns: B (=2) onward.
    for i, t in enumerate(model.time):
        layout.col_index[t] = layout.label_col + 1 + i

    # Row order = topological sort so refs always point upward / leftward.
    import networkx as nx

    from ..solver import build_dag  # local import to avoid cycle

    g = build_dag(model)
    order = [n for n in nx.topological_sort(g) if n in rows]
    # any leftover rows (no edges) get appended in declaration order
    declared = list(rows)
    for n in declared:
        if n not in order:
            order.append(n)

    for i, name in enumerate(order):
        layout.row_index[name] = layout.header_row + 1 + i

    return layout


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------


def _write_globals_sheet(wb: Workbook, model: Model, layout: Layout) -> None:
    ws = wb.create_sheet("globals")
    ws["A1"] = "name"
    ws["B1"] = "value"
    globs = model._globs  # pyright: ignore[reportPrivateUsage]
    for i, (name, _glob) in enumerate(globs.items(), start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=getattr(model, name))
        # Define a workbook-scoped named range pointing at this cell.
        ref = f"'globals'!$B${i}"
        wb.defined_names[name] = DefinedName(name, attr_text=ref)


def _write_scalars_sheet(wb: Workbook, model: Model, layout: Layout) -> None:
    scalars = model._scalars  # pyright: ignore[reportPrivateUsage]
    if not scalars:
        return
    ws = wb.create_sheet("scalars")
    ws["A1"] = "name"
    ws["B1"] = "value"
    for i, name in enumerate(scalars, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=model.cell(name, None))
        ref = f"'scalars'!$B${i}"
        wb.defined_names[name] = DefinedName(name, attr_text=ref)


def _write_datasets(wb: Workbook, model: Model) -> None:
    """Hidden sheets for every CSV dataset attached to the model."""
    for name in dir(type(model)):
        descriptor = getattr(type(model), name, None)
        # The dataset descriptor is a class _CsvDescriptor; check by name to
        # avoid importing it (private).
        if descriptor is None or descriptor.__class__.__name__ != "_CsvDescriptor":
            continue
        try:
            ds = getattr(model, name)
        except Exception:
            continue
        sheet_name = f"ds_{name}"[:31]
        ws = wb.create_sheet(sheet_name)
        ws.sheet_state = "hidden"
        if not ds.rows:
            continue
        cols = ds.columns
        for j, c in enumerate(cols, start=1):
            ws.cell(row=1, column=j, value=c)
        for i, r in enumerate(ds.rows, start=2):
            for j, c in enumerate(cols, start=1):
                ws.cell(row=i, column=j, value=r.get(c))


def _write_model_sheet(wb: Workbook, model: Model, layout: Layout) -> None:
    ws = wb.create_sheet(layout.sheet)
    # Header row.
    ws.cell(row=layout.header_row, column=layout.label_col, value="(label)")
    for t, col in layout.col_index.items():
        ws.cell(row=layout.header_row, column=col, value=t)

    overrides = _override_index(model)
    rows = model._rows  # pyright: ignore[reportPrivateUsage]
    for name, row_obj in rows.items():
        if name not in layout.row_index:
            continue
        ws.cell(row=layout.row_index[name], column=layout.label_col, value=name)
        for t, col in layout.col_index.items():
            target = ws.cell(row=layout.row_index[name], column=col)
            value = model.cell(name, t)
            if (name, t) in overrides:
                # Hardcode override values per PRD §3.7 / §9.2
                target.value = value
                _comment(target, f"Override: {overrides[(name, t)]}")
                continue
            try:
                formula = row_formula_for_period(model, row_obj, t, layout)
                target.value = formula
            except UnsupportedFormula:
                # Fallback: write the raw value plus a comment.
                target.value = value
                _comment(target, "value (formula unsupported in export)")

    # Decent default column width.
    ws.column_dimensions["A"].width = 16
    for t in model.time:
        ws.column_dimensions[get_column_letter(layout.col_index[t])].width = 14


def _override_index(model: Model) -> dict[tuple[str, Any], str]:
    """Return ``{(row, period): reason}`` for any overrides recorded.

    Overrides applied via :func:`model.overrides.apply_overrides` populate
    cells before solve; we don't keep the metadata around. Phase 2 keeps this
    plumbing lightweight: callers pass overrides metadata via
    ``model._exported_overrides`` if they want comments. Empty by default.
    """
    return getattr(model, "_exported_overrides", {})


def _comment(cell: Any, text: str) -> None:
    from openpyxl.comments import Comment

    cell.comment = Comment(text, "model")
