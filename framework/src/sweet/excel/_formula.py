"""Translate a Layer-1 row-body AST into an Excel formula string.

We assume the body is in canonical Layer-1 form (Phase-0 / Phase-1 sugar
already ran). For each period ``t`` we produce the formula that should appear
in the corresponding xlsx cell. The translator handles:

- ``self.budget(t - k)`` → cell ref to ``budget`` at ``t-k`` (inside the row).
- ``self.budget(2024)`` → absolute cell ref at the literal period.
- ``self.budget(self.time.first|.last)`` → cell ref at first/last period.
- ``self.<glob>`` → named range reference.
- ``self.series("<row>")`` → range reference covering the row's entire span.
- numeric/boolean literals, +-*/% and comparisons.
- ``if t == X: return A; ...; return Default`` shape (the canonical Layer-2
  desugar) becomes a chain of Excel ``IF(...)`` calls.

Anything we can't translate cleanly raises :class:`UnsupportedFormula`. The
caller (:mod:`.export`) then writes the *value* into the cell instead, with a
comment explaining the fallback. This keeps round-trip *values* correct even
when an expression is too exotic to render as a formula.
"""

from __future__ import annotations

import ast
import inspect
import textwrap
from typing import Any

from openpyxl.utils import get_column_letter

from ..core import Model, Row
from ._layout import Layout


class UnsupportedFormula(Exception):
    """Raised when a row body can't be rendered as an Excel formula."""


# ---------------------------------------------------------------------------
# Public entry
# ---------------------------------------------------------------------------


def row_formula_for_period(
    model: Model,
    row_obj: Row,
    period: Any,
    layout: Layout,
) -> str:
    """Return the Excel formula string (starting with ``=``) for one cell.

    Falls back to raising :class:`UnsupportedFormula` when the row body uses
    constructs not yet wired up.
    """
    fn = row_obj.fn
    src = textwrap.dedent(inspect.getsource(fn))
    tree = ast.parse(src)
    funcdef = tree.body[0]
    if not isinstance(funcdef, ast.FunctionDef):
        raise UnsupportedFormula(f"row {row_obj.name!r}: not a function definition")
    expr = _select_branch_for_period(funcdef.body, period, model)
    return "=" + _render(expr, model, period, layout, row_obj.name)


# ---------------------------------------------------------------------------
# Body shape: figure out which branch applies for a given period.
# ---------------------------------------------------------------------------


def _select_branch_for_period(
    body: list[ast.stmt], period: Any, model: Model
) -> ast.expr:
    """Walk the canonical Layer-1 body and return the expression for ``period``.

    Recognized shape::

        if t == <const>:
            return <expr>
        if t == <const>:
            return <expr>
        return <default_expr>

    Plain ``return <expr>`` is the all-periods default.
    """
    default: ast.expr | None = None
    for stmt in body:
        if isinstance(stmt, ast.If) and _is_period_eq_test(stmt.test):
            guard_value = _eval_period_constant(stmt.test, model)  # type: ignore[arg-type]
            if guard_value == period:
                return _extract_return(stmt.body)
            continue
        if isinstance(stmt, ast.Return):
            default = stmt.value if stmt.value is not None else None
            continue
        if isinstance(stmt, ast.Expr):
            continue  # docstring etc.
        raise UnsupportedFormula(
            f"unsupported statement in row body: {type(stmt).__name__}"
        )
    if default is None:
        raise UnsupportedFormula("no default return in row body")
    return default


def _is_period_eq_test(test: ast.expr) -> bool:
    return (
        isinstance(test, ast.Compare)
        and len(test.ops) == 1
        and isinstance(test.ops[0], ast.Eq)
        and isinstance(test.left, ast.Name)
        and test.left.id == "t"
    )


def _eval_period_constant(test: ast.Compare, model: Model) -> Any:
    """Evaluate the RHS of ``t == <expr>`` using the model's first/last."""
    rhs = test.comparators[0]
    if isinstance(rhs, ast.Constant):
        return rhs.value
    # self.time.first / self.time.last
    if (
        isinstance(rhs, ast.Attribute)
        and isinstance(rhs.value, ast.Attribute)
        and isinstance(rhs.value.value, ast.Name)
        and rhs.value.value.id == "self"
        and rhs.value.attr == "time"
        and rhs.attr in ("first", "last")
    ):
        return model.time.first if rhs.attr == "first" else model.time.last
    raise UnsupportedFormula(f"unsupported guard expression: {ast.dump(rhs)}")


def _extract_return(stmts: list[ast.stmt]) -> ast.expr:
    if len(stmts) == 1 and isinstance(stmts[0], ast.Return) and stmts[0].value is not None:
        return stmts[0].value
    raise UnsupportedFormula("if branch must contain a single return")


# ---------------------------------------------------------------------------
# Expression renderer
# ---------------------------------------------------------------------------

_BIN_OP = {
    ast.Add: "+",
    ast.Sub: "-",
    ast.Mult: "*",
    ast.Div: "/",
    ast.Mod: "MOD",  # special-cased
    ast.Pow: "^",
    ast.FloorDiv: "INT(",  # special-cased
}

_CMP_OP = {
    ast.Eq: "=",
    ast.NotEq: "<>",
    ast.Lt: "<",
    ast.LtE: "<=",
    ast.Gt: ">",
    ast.GtE: ">=",
}


def _render(
    node: ast.expr,
    model: Model,
    period: Any,
    layout: Layout,
    self_row: str,
) -> str:
    if isinstance(node, ast.Constant):
        v = node.value
        if isinstance(v, bool):
            return "TRUE" if v else "FALSE"
        if isinstance(v, (int, float)):
            return repr(v)
        if isinstance(v, str):
            escaped = v.replace('"', '""')
            return f'"{escaped}"'
        raise UnsupportedFormula(f"unsupported constant {v!r}")

    if isinstance(node, ast.UnaryOp):
        if isinstance(node.op, ast.USub):
            return f"(-{_render(node.operand, model, period, layout, self_row)})"
        if isinstance(node.op, ast.UAdd):
            return _render(node.operand, model, period, layout, self_row)
        if isinstance(node.op, ast.Not):
            return f"NOT({_render(node.operand, model, period, layout, self_row)})"
        raise UnsupportedFormula(f"unsupported unary op {type(node.op).__name__}")

    if isinstance(node, ast.BinOp):
        op_t = type(node.op)
        if op_t is ast.Mod:
            lhs = _render(node.left, model, period, layout, self_row)
            rhs = _render(node.right, model, period, layout, self_row)
            return f"MOD({lhs},{rhs})"
        if op_t is ast.FloorDiv:
            lhs = _render(node.left, model, period, layout, self_row)
            rhs = _render(node.right, model, period, layout, self_row)
            return f"INT(({lhs})/({rhs}))"
        if op_t not in _BIN_OP:
            raise UnsupportedFormula(f"unsupported binary op {op_t.__name__}")
        op = _BIN_OP[op_t]
        lhs = _render(node.left, model, period, layout, self_row)
        rhs = _render(node.right, model, period, layout, self_row)
        return f"({lhs}{op}{rhs})"

    if isinstance(node, ast.BoolOp):
        op = "AND" if isinstance(node.op, ast.And) else "OR"
        rendered = [_render(v, model, period, layout, self_row) for v in node.values]
        return f"{op}({','.join(rendered)})"

    if isinstance(node, ast.Compare):
        if len(node.ops) != 1:
            raise UnsupportedFormula("chained comparisons not supported")
        op = _CMP_OP.get(type(node.ops[0]))
        if op is None:
            raise UnsupportedFormula(f"unsupported compare op {type(node.ops[0]).__name__}")
        lhs = _render(node.left, model, period, layout, self_row)
        rhs = _render(node.comparators[0], model, period, layout, self_row)
        return f"({lhs}{op}{rhs})"

    if isinstance(node, ast.Attribute):
        return _render_attribute(node, model, layout)

    if isinstance(node, ast.Name):
        if node.id == "t":
            return repr(period)
        raise UnsupportedFormula(f"bare name {node.id!r} in expression")

    if isinstance(node, ast.Call):
        return _render_call(node, model, period, layout, self_row)

    raise UnsupportedFormula(f"unsupported expression node {type(node).__name__}")


def _render_attribute(node: ast.Attribute, model: Model, layout: Layout) -> str:
    # self.<glob>  → named range (just the name)
    # self.time.first / .last → period literal
    if isinstance(node.value, ast.Name) and node.value.id == "self":
        attr = node.attr
        if attr in layout.glob_names or attr in layout.scalar_names:
            return attr
        raise UnsupportedFormula(f"unsupported attribute self.{attr}")
    if (
        isinstance(node.value, ast.Attribute)
        and isinstance(node.value.value, ast.Name)
        and node.value.value.id == "self"
        and node.value.attr == "time"
        and node.attr in ("first", "last")
    ):
        return repr(model.time.first if node.attr == "first" else model.time.last)
    raise UnsupportedFormula(f"unsupported attribute access {ast.dump(node)}")


def _render_call(
    node: ast.Call, model: Model, period: Any, layout: Layout, self_row: str
) -> str:
    # self.<row>(<period_expr>) → cell ref for that period
    if (
        isinstance(node.func, ast.Attribute)
        and isinstance(node.func.value, ast.Name)
        and node.func.value.id == "self"
        and node.func.attr in layout.row_names
    ):
        row_name = node.func.attr
        if len(node.args) != 1:
            raise UnsupportedFormula(f"row call {row_name} expects 1 arg, got {len(node.args)}")
        target_period = _eval_period_arg(node.args[0], model, period)
        return _make_cell_ref(layout, row_name, target_period, same_row=row_name == self_row)
    # self.series("<row>") → full-row range
    if (
        isinstance(node.func, ast.Attribute)
        and isinstance(node.func.value, ast.Name)
        and node.func.value.id == "self"
        and node.func.attr == "series"
        and len(node.args) == 1
        and isinstance(node.args[0], ast.Constant)
        and isinstance(node.args[0].value, str)
    ):
        row_name = node.args[0].value
        return _make_row_range(layout, row_name, model)
    raise UnsupportedFormula(f"unsupported call {ast.dump(node.func)}")


def _eval_period_arg(arg: ast.expr, model: Model, period: Any) -> Any:
    """Evaluate a period expression statically (no runtime, just AST)."""
    if isinstance(arg, ast.Constant):
        return arg.value
    if isinstance(arg, ast.Name) and arg.id == "t":
        return period
    if isinstance(arg, ast.BinOp) and isinstance(arg.op, (ast.Add, ast.Sub)):
        lhs = _eval_period_arg(arg.left, model, period)
        rhs = _eval_period_arg(arg.right, model, period)
        return lhs + rhs if isinstance(arg.op, ast.Add) else lhs - rhs
    if (
        isinstance(arg, ast.Attribute)
        and isinstance(arg.value, ast.Attribute)
        and isinstance(arg.value.value, ast.Name)
        and arg.value.value.id == "self"
        and arg.value.attr == "time"
        and arg.attr in ("first", "last")
    ):
        return model.time.first if arg.attr == "first" else model.time.last
    raise UnsupportedFormula(f"can't evaluate period expr: {ast.dump(arg)}")


def _make_cell_ref(layout: Layout, row_name: str, period: Any, *, same_row: bool) -> str:
    if period not in layout.col_index:
        raise UnsupportedFormula(f"period {period!r} out of model time range")
    col = get_column_letter(layout.col_index[period])
    r = layout.row_index[row_name]
    # Self-references inside the same row use a relative ref (no $) so the
    # formula can be copied across columns naturally; cross-row refs use an
    # absolute row + relative col so they "follow t" too.
    if same_row:
        return f"{col}{r}"
    return f"{col}${r}"


def _make_row_range(layout: Layout, row_name: str, model: Model) -> str:
    first_col = get_column_letter(layout.col_index[model.time.first])
    last_col = get_column_letter(layout.col_index[model.time.last])
    r = layout.row_index[row_name]
    return f"{first_col}${r}:{last_col}${r}"
