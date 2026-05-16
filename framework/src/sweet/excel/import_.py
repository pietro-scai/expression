"""``sweet import`` — ``.xlsx`` → ``sweet.py`` + ``sweet.md`` skeleton (PRD §9.1).

Phase 2 ships a *non-interactive* importer that produces a Layer-1 skeleton
and a ``sweet.md`` with TODOs for ambiguities. Agent dialogue (PRD §9.1 step
4 — confirmation prompts) is wired up in Phase 3.

Heuristics:
- Time axis: if the first row is a sequence of consecutive integers (years),
  treat that as ``periods(start, end)``.
- Globals: any single cell with a defined name becomes a glob with that name
  and the cell's value as default.
- Rows: any row whose label is a non-empty string in column A and whose
  data cells are formulas (or numbers) becomes a ``@row``. We emit Layer-1
  bodies that *encode the values* — preserving solve correctness — with a
  TODO marker prompting the user to replace with the proper formula.

The contract is that ``sweet run`` after import reproduces every cell value
of the imported workbook. Higher-fidelity formula re-derivation is a Phase
3 effort with the agent in the loop.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class ImportPlan:
    """The parsed structure ready to be rendered as Python source."""

    classname: str
    period_start: int
    period_end: int
    globs: list[tuple[str, Any]] = field(default_factory=list)  # (name, default)
    rows: list[tuple[str, list[tuple[Any, Any]]]] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)
    sheet_used: str = "model"


def import_xlsx(path: str | Path, classname: str = "Imported") -> ImportPlan:
    """Parse ``path`` into an :class:`ImportPlan` (no I/O on the way out)."""
    p = Path(path)
    wb = load_workbook(p, data_only=True)
    ws = _pick_sheet(wb)

    period_start, period_end, period_cols, time_issues = _detect_time_axis(ws)
    globs = _detect_globs(wb)

    rows: list[tuple[str, list[tuple[Any, Any]]]] = []
    issues: list[str] = list(time_issues)

    if period_start is None:
        issues.append("Could not detect a time axis (looking for an integer header row).")

    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if not isinstance(label, str) or not label.strip():
            continue
        clean = _slugify(label.strip())
        if not clean:
            continue
        cells: list[tuple[Any, Any]] = []
        for period, col in period_cols.items():
            v = ws.cell(row=row_idx, column=col).value
            if v is None:
                continue
            cells.append((period, v))
        if not cells:
            continue
        if clean != label.strip():
            issues.append(
                f"Row label {label!r} renamed to {clean!r} (Python identifier rules)."
            )
        rows.append((clean, cells))

    return ImportPlan(
        classname=classname,
        period_start=period_start or 0,
        period_end=period_end or 0,
        globs=globs,
        rows=rows,
        issues=issues,
        sheet_used=ws.title,
    )


def render_plan(plan: ImportPlan) -> tuple[str, str]:
    """Return ``(model_py_text, model_md_text)`` for the import."""
    return _render_python(plan), _render_markdown(plan)


def write_imported(plan: ImportPlan, dest: Path) -> None:
    """Write ``sweet.py`` + ``sweet.md`` (and the standard scaffold dirs) to ``dest``."""
    dest.mkdir(parents=True, exist_ok=True)
    (dest / "inputs").mkdir(exist_ok=True)
    (dest / "outputs").mkdir(exist_ok=True)
    (dest / "tests").mkdir(exist_ok=True)
    py, md = render_plan(plan)
    (dest / "sweet.py").write_text(py)
    (dest / "sweet.md").write_text(md)


# ---------------------------------------------------------------------------
# Heuristics
# ---------------------------------------------------------------------------


def _pick_sheet(wb: Any) -> Any:
    if "model" in wb.sheetnames:
        return wb["model"]
    return wb.active


def _detect_time_axis(ws: Any) -> tuple[int | None, int | None, dict[int, int], list[str]]:
    """Look at row 1 for an integer year sequence.

    Returns ``(start, end, {period: col_index}, issues)``.
    """
    issues: list[str] = []
    period_cols: dict[int, int] = {}
    seen: list[int] = []
    for col_idx in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=col_idx).value
        if isinstance(v, int):
            period_cols[v] = col_idx
            seen.append(v)
    if not seen:
        return None, None, {}, issues
    sorted_seen = sorted(seen)
    if sorted_seen != list(range(sorted_seen[0], sorted_seen[-1] + 1)):
        issues.append(
            "Time axis is non-contiguous: "
            f"{sorted_seen}. Phase 2 requires consecutive periods."
        )
    return sorted_seen[0], sorted_seen[-1], period_cols, issues


def _detect_globs(wb: Any) -> list[tuple[str, Any]]:
    """Named ranges that point at a single cell become globals."""
    out: list[tuple[str, Any]] = []
    for name, dn in wb.defined_names.items():
        ref = dn.attr_text  # e.g. "'globals'!$B$2"
        m = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", ref)
        if not m:
            continue
        sheet, col, row = m.group(1), m.group(2), int(m.group(3))
        if sheet not in wb.sheetnames:
            continue
        v = wb[sheet][f"{col}{row}"].value
        if v is None:
            continue
        if not _is_identifier(name):
            continue
        out.append((name, v))
    return out


_IDENT_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _is_identifier(s: str) -> bool:
    return bool(_IDENT_RE.match(s))


def _slugify(s: str) -> str:
    """Best-effort label → Python identifier."""
    out = re.sub(r"[^A-Za-z0-9_]+", "_", s).strip("_")
    if out and out[0].isdigit():
        out = "row_" + out
    return out.lower()


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------


def _render_python(plan: ImportPlan) -> str:
    lines: list[str] = []
    lines.append('"""Generated by ``sweet import`` — Phase 2 (Layer 1)."""')
    lines.append("")
    lines.append("from sweet import Model, glob, periods, row")
    lines.append("")
    lines.append("")
    lines.append(f"class {plan.classname}(Model):")
    lines.append(f"    time = periods({plan.period_start}, {plan.period_end})")
    if plan.globs:
        lines.append("")
        for name, val in plan.globs:
            lines.append(f"    {name} = glob({val!r})")
    for name, cells in plan.rows:
        lines.append("")
        lines.append("    @row")
        lines.append(f"    def {name}(self, t):")
        if not cells:
            lines.append("        return None  # TODO: empty row in source")
            continue
        emitted_default = False
        for i, (period, value) in enumerate(cells):
            guard = f"t == {period}"
            prefix = "if" if i == 0 else "elif"
            lines.append(f"        {prefix} {guard}:")
            lines.append(f"            return {_py_literal(value)}")
            emitted_default = True
        if emitted_default:
            lines.append("        # TODO: replace literal values above with the formula(s) "
                         "from the source Excel.")
            lines.append('        raise ValueError(f"no rule for t={t}")')
    lines.append("")
    return "\n".join(lines)


def _py_literal(v: Any) -> str:
    if isinstance(v, str) and v.startswith("="):
        # The source had a formula; we don't carry it across — use the
        # cached value if any, else encode the formula text as a comment.
        return f"None  # source formula: {v!r}"
    if isinstance(v, str):
        return repr(v)
    return repr(v)


def _render_markdown(plan: ImportPlan) -> str:
    md: list[str] = []
    md.append(f"# {plan.classname} model")
    md.append("")
    md.append("## Purpose")
    md.append("")
    md.append("> One paragraph: what business question does this answer? "
              "(filled in during agent dialogue)")
    md.append("")
    md.append("## Inputs")
    md.append("")
    if plan.globs:
        for name, val in plan.globs:
            md.append(f"- `{name}` (global, default `{val!r}`)")
    else:
        md.append("- (no named globals detected)")
    md.append("")
    md.append("## Outputs")
    md.append("")
    if plan.rows:
        for name, _cells in plan.rows:
            md.append(f"- `{name}[year]`")
    else:
        md.append("- (no row outputs detected)")
    md.append("")
    md.append("## Logic")
    md.append("")
    md.append("> TODO: explain the model in plain prose. The auto-generated "
              "row bodies preserve the *values* from the source Excel but "
              "encode them as per-period literals — replace each `if/elif` "
              "with the actual formula from the source.")
    md.append("")
    md.append("## Issues found during import")
    md.append("")
    if plan.issues:
        for issue in plan.issues:
            md.append(f"- [ ] {issue}")
    else:
        md.append("- [x] None")
    md.append("")
    md.append("## Overrides")
    md.append("")
    md.append("None recorded.")
    md.append("")
    md.append("## Changelog")
    md.append("")
    md.append("- imported via `model import`.")
    md.append("")
    return "\n".join(md)
