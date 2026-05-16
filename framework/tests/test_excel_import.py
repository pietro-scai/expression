"""Phase 2: Excel import tests."""

from __future__ import annotations

import importlib.util
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from sweet.excel import export, import_xlsx
from sweet.excel.import_ import render_plan, write_imported


def _build_simple_workbook(path: Path) -> None:
    """Hand-build a minimal model-shaped workbook."""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    g = wb.create_sheet("globals")
    g["A1"] = "name"
    g["B1"] = "value"
    g["A2"] = "rate"
    g["B2"] = 0.05
    wb.defined_names["rate"] = DefinedName("rate", attr_text="'globals'!$B$2")

    ws = wb.create_sheet("model")
    ws["A1"] = "(label)"
    for col, year in enumerate([2024, 2025, 2026], start=2):
        ws.cell(row=1, column=col, value=year)
    ws["A2"] = "budget"
    ws["B2"] = 100
    ws["C2"] = 105
    ws["D2"] = 110.25
    wb.save(path)


def test_import_detects_time_axis(tmp_path: Path):
    src = tmp_path / "src.xlsx"
    _build_simple_workbook(src)
    plan = import_xlsx(src, classname="Imported")
    assert plan.period_start == 2024
    assert plan.period_end == 2026


def test_import_detects_globals(tmp_path: Path):
    src = tmp_path / "src.xlsx"
    _build_simple_workbook(src)
    plan = import_xlsx(src)
    names = [n for n, _ in plan.globs]
    assert "rate" in names


def test_import_detects_rows(tmp_path: Path):
    src = tmp_path / "src.xlsx"
    _build_simple_workbook(src)
    plan = import_xlsx(src)
    row_names = [n for n, _ in plan.rows]
    assert "budget" in row_names


def test_import_renders_python(tmp_path: Path):
    src = tmp_path / "src.xlsx"
    _build_simple_workbook(src)
    plan = import_xlsx(src, classname="Test")
    py, md = render_plan(plan)
    assert "class Test(Model):" in py
    assert "time = periods(2024, 2026)" in py
    assert "rate = glob(0.05)" in py
    assert "def budget" in py
    assert "# Test model" in md
    assert "## Issues found" in md


def test_imported_model_solves(tmp_path: Path):
    src = tmp_path / "src.xlsx"
    _build_simple_workbook(src)
    plan = import_xlsx(src, classname="Imported")
    dest = tmp_path / "demo"
    write_imported(plan, dest)

    # Load the generated model.py and verify it solves.
    spec = importlib.util.spec_from_file_location("imported_test_module", dest / "sweet.py")
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules["imported_test_module"] = module
    spec.loader.exec_module(module)

    cls = module.Imported
    inst = cls().solve()
    assert inst.cell("budget", 2024) == 100
    assert inst.cell("budget", 2025) == 105
    assert inst.cell("budget", 2026) == 110.25


def test_import_then_export_roundtrip(tmp_path: Path):
    """The full PRD §9.1 contract: import → solve → export → verify."""
    src = tmp_path / "src.xlsx"
    _build_simple_workbook(src)
    plan = import_xlsx(src, classname="Imported")
    dest = tmp_path / "demo"
    write_imported(plan, dest)

    spec = importlib.util.spec_from_file_location("imported_rt", dest / "sweet.py")
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules["imported_rt"] = module
    spec.loader.exec_module(module)
    inst = module.Imported().solve()

    out = export(inst, dest / "outputs" / "model.xlsx")
    assert out.exists()


def test_import_handles_missing_time_axis(tmp_path: Path):
    """A workbook with no integer header row should produce an issue."""
    src = tmp_path / "no_time.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "model"
    ws["A1"] = "(label)"
    ws["A2"] = "budget"
    ws["B2"] = 100
    wb.save(src)

    plan = import_xlsx(src)
    assert plan.period_start == 0
    assert any("time axis" in i.lower() for i in plan.issues)


# ---------------------------------------------------------------------------
# CLI smoke
# ---------------------------------------------------------------------------


def _model_bin() -> str:
    bin_path = Path(sys.executable).parent / "sweet"
    if not bin_path.exists():
        pytest.skip(f"Console script not installed at {bin_path}")
    return str(bin_path)


def test_cli_export_and_import(tmp_path: Path):
    # 1. Generate a model with `model init`, run it, export, then re-import.
    proc = subprocess.run(
        [_model_bin(), "init", "demo"],
        cwd=tmp_path, capture_output=True, text=True, check=False,
    )
    assert proc.returncode == 0, proc.stderr
    demo = tmp_path / "demo"

    proc = subprocess.run(
        [_model_bin(), "export", "--out", "outputs/budget.xlsx"],
        cwd=demo, capture_output=True, text=True, check=False,
    )
    assert proc.returncode == 0, proc.stderr
    assert "Exported" in proc.stdout
    assert "Verified" in proc.stdout
    assert (demo / "outputs" / "budget.xlsx").exists()

    proc = subprocess.run(
        [_model_bin(), "import", str(demo / "outputs" / "budget.xlsx"),
         "--out", str(tmp_path / "reimported"), "--classname", "Reimported"],
        capture_output=True, text=True, check=False,
    )
    assert proc.returncode == 0, proc.stderr
    assert (tmp_path / "reimported" / "sweet.py").exists()
    assert (tmp_path / "reimported" / "sweet.md").exists()
