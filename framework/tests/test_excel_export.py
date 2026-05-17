"""Phase 2: Excel export + round-trip verification tests."""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import load_workbook

from expression import Model, glob, periods, row
from expression.excel import export, verify

# ---------------------------------------------------------------------------
# PRD example 12.1 — round trip
# ---------------------------------------------------------------------------


class Budget(Model):
    time = periods(2024, 2028)
    seed = glob(100, doc="Starting budget in $K")
    growth_rate = glob(0.05, doc="Annual growth rate")

    @row
    def budget(self, t):
        if t == self.time.first:
            return self.seed
        return self.budget(t - 1) * (1 + self.growth_rate)


def test_export_writes_xlsx(tmp_path: Path):
    m = Budget().solve()
    out = export(m, tmp_path / "budget.xlsx")
    assert out.exists()
    wb = load_workbook(out, data_only=False)
    assert "model" in wb.sheetnames
    assert "globals" in wb.sheetnames

    ws = wb["model"]
    # Header row: A1 label, B1..F1 = 2024..2028.
    assert ws["A1"].value == "(label)"
    assert [ws.cell(row=1, column=c).value for c in range(2, 7)] == [2024, 2025, 2026, 2027, 2028]
    # Row 2 = budget.
    assert ws["A2"].value == "budget"
    # First-period cell is the seed via named range.
    first = ws["B2"].value
    assert isinstance(first, str) and first.startswith("=")
    assert "seed" in first
    # Recurrence: cells C2..F2 reference the previous column.
    for col in ("C", "D", "E", "F"):
        v = ws[f"{col}2"].value
        assert isinstance(v, str) and v.startswith("=")
        assert "growth_rate" in v


def test_export_globals_named_ranges(tmp_path: Path):
    m = Budget().solve()
    out = export(m, tmp_path / "budget.xlsx")
    wb = load_workbook(out, data_only=False)
    assert "seed" in wb.defined_names
    assert "growth_rate" in wb.defined_names


def test_verify_roundtrip_budget(tmp_path: Path):
    m = Budget().solve()
    out = export(m, tmp_path / "budget.xlsx")
    result = verify(m, out)
    assert result.ok, result.first_mismatches()
    assert result.checked == 5  # 5 budget cells


# ---------------------------------------------------------------------------
# Multi-row PnL (PRD example 12.2 in Layer 1)
# ---------------------------------------------------------------------------


class PnL(Model):
    time = periods(2024, 2026)
    revenue_growth = glob(0.10)
    cogs_pct = glob(0.40)

    @row
    def revenue(self, t):
        if t == self.time.first:
            return 1000
        return self.revenue(t - 1) * (1 + self.revenue_growth)

    @row
    def cogs(self, t):
        return self.revenue(t) * self.cogs_pct

    @row
    def gross_profit(self, t):
        return self.revenue(t) - self.cogs(t)


def test_verify_roundtrip_pnl(tmp_path: Path):
    m = PnL().solve()
    out = export(m, tmp_path / "pnl.xlsx")
    result = verify(m, out)
    assert result.ok, result.first_mismatches()
    # 3 rows x 3 periods = 9 cells.
    assert result.checked == 9


# ---------------------------------------------------------------------------
# Layer-2 form roundtrips too (sugar produces Layer-1 internally).
# ---------------------------------------------------------------------------


class BudgetSugar(Model):
    time = periods(2024, 2026)
    seed = glob(100)
    growth_rate = glob(0.05)

    @row
    def budget():  # type: ignore[no-untyped-def]
        budget[first] = seed  # noqa: F821
        budget[n] = budget[n - 1] * (1 + growth_rate)  # noqa: F821


def test_verify_roundtrip_layer2(tmp_path: Path):
    m = BudgetSugar().solve()
    out = export(m, tmp_path / "sugar.xlsx")
    result = verify(m, out)
    assert result.ok, result.first_mismatches()


# ---------------------------------------------------------------------------
# Verifier reports mismatch when the file is hand-edited.
# ---------------------------------------------------------------------------


def test_verifier_catches_corruption(tmp_path: Path):
    m = Budget().solve()
    out = export(m, tmp_path / "budget.xlsx")
    wb = load_workbook(out)
    wb["model"]["B2"] = 999.0  # stomp seed cell
    wb.save(out)
    result = verify(m, out)
    assert not result.ok
    assert any(mm.cell == "B2" for mm in result.mismatches)


# ---------------------------------------------------------------------------
# Constants as scalars (e.g., a model with only @scalar rows)
# ---------------------------------------------------------------------------


def test_export_handles_empty_globals(tmp_path: Path):
    class M(Model):
        time = periods(2024, 2025)

        @row
        def x(self, t):
            return 42

    m = M().solve()
    out = export(m, tmp_path / "m.xlsx")
    result = verify(m, out)
    assert result.ok


def test_export_skip_verify_returns(tmp_path: Path):
    """Exporting without verification still produces a file."""
    m = Budget().solve()
    out = export(m, tmp_path / "skip.xlsx")
    assert out.exists()


# ---------------------------------------------------------------------------
# Tolerance
# ---------------------------------------------------------------------------


def test_verifier_tolerance_default():
    """Default is 1e-9 relative — small float drift should pass."""
    a = 100.0
    b = 100.0 * (1 + 1e-12)
    assert math.isclose(a, b, rel_tol=1e-9, abs_tol=1e-9)
